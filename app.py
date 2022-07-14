import openpyxl as xl
from openpyxl.chart import (
    BarChart,
    Reference
)

wb = xl.load_workbook('transactions.xlsx');
sheet = wb['Sheet1']  ##acess sheet

cell_model1=sheet['b1'];
cell =sheet.cell(1,1);
print(cell.value, cell_model1.value );

print(sheet.max_row);

for row in range(2,sheet.max_row+1):
    corrected_value=sheet.cell(row,3).value*0.9
    print(corrected_value);
    sheet.cell(row,4).value=corrected_value     #adding a new column for corrected value
    
values=Reference(sheet,min_row=2,max_row=4,min_col=4,max_col=4)
chartObj = BarChart()
chartObj.add_data(values)
sheet.add_chart(chartObj)
wb.save('transactions.xlsx');